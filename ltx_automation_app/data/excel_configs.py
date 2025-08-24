# ltx_automation_app/data/excel_configs.py

# ============ EXCEL FORMATTING CONFIGURATIONS ============
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ============ COLOR DEFINITIONS ============
COLORS = {
    "purple_header": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    "light_purple": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
    "light_blue": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "yellow": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    "green": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "orange": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
}

# ============ BORDER STYLES ============
BORDERS = {
    "thin": Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    ),
    "thick": Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    ),
    "thick_outline": Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
}

# ============ FONT STYLES ============
FONTS = {
    "header_white": Font(name='Calibri', bold=True, size=18, color="FFFFFF"),
    "header_black": Font(name='Calibri', bold=True, size=16, color="000000"),
    "subheader": Font(name='Calibri', bold=True, size=14, color="000000"),
    "normal": Font(name='Calibri', size=14, color="000000"),
    "small": Font(name='Calibri', size=12, color="000000"),
    "evergreen_label": Font(name='Calibri', size=14, color="00A500", bold=True),
    "custom_label": Font(name='Calibri', size=14, color="D87A00", bold=True),
    "blue_text": Font(name='Calibri', size=14, color="0070C0", bold=True)
}

# ============ COLUMN WIDTHS ============
COLUMN_WIDTHS = {
    "readme": {
        'A': 2.00,
        'B': 15.00, 'C': 15.00, 'D': 15.00, 'E': 15.00,
        'F': 10.00, 'G': 10.00, 'H': 10.00, 'I': 10.00,
        'J': 15.00, 'K': 15.00, 'L': 15.00,
        'M': 10.00,
        'N': 15.00, 'O': 15.00, 'P': 15.00, 'Q': 15.00
    },
    "part1": {
        'A': 8,      # TYPE
        'B': 50,     # SOURCE
        'C': 50,     # TARGET
        'D': 11,     # Word Count
        'E': 20,     # Pre-Eval
        'F': 20,     # Applicable Word Count 1
        'G': 25,     # Applicable Word Count 2
        'H': 10,     # Overall
        'I': 10,     # Accuracy
        'J': 17,     # Omission/Addition
        'K': 12,     # Compliance
        'L': 10,     # Fluency
        'M': 15,     # Rating (Not Weighted)
        'N': 15,     # Rating (Weighted)
        'O': 10,     # Tag/URL (custom metric)
        'P': 30      # Additional notes
    },
    "part2": {
        "default": 10  # All columns width 10
    },
    "part3": {
        "default": 12  # All columns width 12
    }
}

# ============ ROW HEIGHTS ============
ROW_HEIGHTS = {
    "header": 85,
    "instruction": 20,
    "metric_header": 35,
    "default": 15
}

# ============ MERGE PATTERNS ============
MERGE_PATTERNS = {
    "readme": {
        "instructions": "B{start}:Q{end}",
        "stakeholder": "B{start}:Q{end}",
        "metric_headers": {
            "category": "B{row}:C{row}",
            "metrics": "D{row}:E{row}",
            "definitions": "F{row}:I{row}",
            "notes": "J{row}:L{row}",
            "weights": "M{row}:M{row}",
            "weight_def": "N{row}:O{row}",
            "scoring_def": "P{row}:Q{row}"
        }
    }
}

# ============ FORMULAS ============
FORMULAS = {
    "word_count": '=IF(B{row}="","",LEN(B{row})-LEN(SUBSTITUTE(B{row}," ",""))+1)',
    "applicable_count_1": '=IF(E{row}="Incomprehensible Input","-",D{row})',
    "applicable_count_2": '=IF(E{row}="Irrelevant Output",D{row},IF(E{row}="",D{row},"-"))',
    "rating_not_weighted": '=IF(COUNT(I{row}:L{row})=0,"",AVERAGE(I{row}:L{row}))',
    "rating_weighted": (
        '=IF(COUNT(I{row}:L{row})=0,"",'
        '(I{row}*FORMULA_HELPER!$B$2+'
        'J{row}*FORMULA_HELPER!$B$3+'
        'K{row}*FORMULA_HELPER!$B$4+'
        'L{row}*FORMULA_HELPER!$B$5+'
        'IF(ISNUMBER(O{row}),O{row}*FORMULA_HELPER!$E$2,0))'
        '/FORMULA_HELPER!$G$2)'
    ),
    "weight_sum": "=SUM(M{start}:M{end})"
}

# ============ VALIDATION RULES ============
VALIDATIONS = {
    "pre_eval": {
        "type": "list",
        "formula1": '"Incomprehensible Input,Irrelevant Output"',
        "allow_blank": True,
        "error": "Please select from the list",
        "error_title": "Invalid Entry"
    },
    "score": {
        "type": "whole",
        "operator": "between",
        "formula1": 1,
        "formula2": 5,
        "allow_blank": True,
        "error": "Score must be between 1 and 5",
        "error_title": "Invalid Score"
    }
}

# ============ SHEET CONFIGURATIONS ============
SHEET_CONFIGS = {
    "readme": {
        "name": "READ ME",
        "tab_color": "FFFFFF",
        "show_gridlines": False,
        "default_row_height": 15
    },
    "formula_helper": {
        "name": "FORMULA_HELPER",
        "hidden": True,
        "tab_color": "FFFFFF",
        "show_gridlines": True
    },
    "part1_model": {
        "name_template": "PART 1 - MODEL {letter}",
        "tab_color": "FFFFFF", 
        "show_gridlines": False,
        "freeze_panes": "A2"
    },
    "part2": {
        "name": "PART 2 - DATA ANALYSIS",
        "tab_color": "FFFFFF",
        "show_gridlines": False
    },
    "part3": {
        "name": "PART 3 - CRITERIA BASED ASSESS",
        "tab_color": "FFFFFF",
        "show_gridlines": False
    }
}

# ============ CONDITIONAL FORMATTING ============
CONDITIONAL_FORMATS = {
    "missing_score": {
        "type": "blanks",
        "format": {"fill": COLORS["yellow"]},
        "applies_to": ["H", "I", "J", "K", "L", "O"]  # Score columns
    }
}