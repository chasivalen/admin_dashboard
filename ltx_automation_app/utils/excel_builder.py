# ltx_automation_app/utils/excel_builder.py
"""
Excel builder that works with database records instead of static files
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from typing import List, Dict, Any, Optional
import base64
import io
import logging

# Keep excel_configs for formatting - this is fine
from ltx_automation_app.data.excel_configs import (
    COLORS, BORDERS, FONTS, COLUMN_WIDTHS, ROW_HEIGHTS, 
    FORMULAS, VALIDATIONS, SHEET_CONFIGS
)

# NO MORE IMPORTS FROM metrics_catalog or readme_templates!
# Data comes from database via the state

# Configure logging
logger = logging.getLogger(__name__)


class DynamicExcelBuilder:
    """
    Builds Excel evaluation templates dynamically based on database data.
    Handles all Excel generation logic with proper error handling.
    """
    
    def __init__(self, state):
        """
        Initialize with state containing database records and user selections.
        
        Args:
            state: State instance with database data and configuration
        """
        self.state = state
        self.wb = Workbook()
        self.current_readme_row = 1
        self.weight_cells = []  # Track cells for weight sum formula
        
    def build(self) -> bytes:
        """
        Build complete Excel workbook and return as bytes.
        
        Returns:
            bytes: Excel file data ready for download
            
        Raises:
            Exception: If Excel generation fails
        """
        try:
            # Remove default sheet
            if self.wb.active:
                self.wb.remove(self.wb.active)
            
            # Create all sheets in order
            self._create_readme_sheet()
            self._create_formula_helper()
            self._create_part1_sheets()
            self._create_part2_sheet()
            self._create_part3_sheet()
            
            # Convert to bytes
            return self._save_to_bytes()
            
        except Exception as e:
            logger.error(f"Failed to build Excel template: {str(e)}")
            raise Exception(f"Excel generation failed: {str(e)}")
    
    # ============ README SHEET BUILDER ============
    def _create_readme_sheet(self):
        """Create the README sheet with instructions and metrics configuration."""
        try:
            ws = self.wb.create_sheet("READ ME")
            self._apply_sheet_config(ws, "readme")
            
            # Place header
            self._place_readme_header(ws)
            
            # Place instructions section
            self._place_instructions_section(ws)
            
            # Place stakeholder section
            self._place_stakeholder_section(ws)
            
            # Place metrics table
            self._place_metrics_table(ws)
            
            # Place total sum
            self._place_metrics_total(ws)
            
            # Place scoring definitions
            self._place_scoring_definitions(ws)
            
        except Exception as e:
            logger.error(f"Failed to create README sheet: {str(e)}")
            raise
    
    def _place_readme_header(self, ws):
        """Place the README header."""
        ws['B1'] = "Instructions for Use:"
        ws['B1'].font = FONTS["header_black"]
        self.current_readme_row = 3
    
    def _place_instructions_section(self, ws):
        """Place instruction lines from database."""
        instructions_start = self.current_readme_row
        
        # Get instruction lines from state (which loaded from database)
        if hasattr(self.state, 'selected_readme') and self.state.selected_readme:
            # Use the selected README from database
            readme_text = self.state.selected_readme.README_TXT
            # Split into lines if it's a multi-line text
            lines = readme_text.split('\n') if readme_text else []
        elif hasattr(self.state, 'custom_readme_lines'):
            # Use custom lines if provided
            lines = self.state.custom_readme_lines
        else:
            # Default empty
            lines = []
        
        # Place each instruction line
        for line in lines:
            # Replace terminology placeholders if they exist
            if hasattr(self.state, 'terminology_choices'):
                if "{source_issue}" in line or "{target_issue}" in line or "{scoring_instruction}" in line:
                    try:
                        processed_line = line.format(
                            source_issue=self.state.terminology_choices.get("source_issue", ""),
                            target_issue=self.state.terminology_choices.get("target_issue", ""),
                            scoring_instruction=self.state.terminology_choices.get("scoring_instruction", "")
                        )
                    except KeyError:
                        processed_line = line  # Use original if formatting fails
                else:
                    processed_line = line
            else:
                processed_line = line
            
            ws[f'B{self.current_readme_row}'] = processed_line
            ws[f'B{self.current_readme_row}'].font = Font(name='Calibri', size=14)
            ws[f'B{self.current_readme_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Merge cells for this row
            ws.merge_cells(f'B{self.current_readme_row}:Q{self.current_readme_row}')
            
            # Set row height
            ws.row_dimensions[self.current_readme_row].height = ROW_HEIGHTS.get("instruction", 20)
            
            self.current_readme_row += 1
        
        # Apply border around entire instruction block
        if self.current_readme_row > instructions_start:
            self._apply_block_border(ws, instructions_start, self.current_readme_row - 1, 2, 17)
    
    def _place_metrics_table(self, ws):
        """Place the metrics table with selected metrics from database."""
        # Table headers
        self._place_metrics_headers(ws)
        
        # Track starting row for weight sum formula
        metrics_start_row = self.current_readme_row
        self.weight_cells = []
        
        # Get metrics from state (loaded from database)
        evergreen_metrics = getattr(self.state, 'evergreen_metrics_db', [])
        custom_metrics = getattr(self.state, 'custom_metrics_db', [])
        
        # Place selected metrics
        evergreen_start = None
        custom_start = None
        
        # Add evergreen metrics from database
        for metric in evergreen_metrics:
            if evergreen_start is None:
                evergreen_start = self.current_readme_row
            
            self._place_metric_row_from_db(ws, metric)
            
            # Track weight cell if metric has weight
            weight = getattr(metric, 'METRIC_WEIGHT', None) or getattr(metric, 'weight', 5)
            if weight:
                self.weight_cells.append(f'M{self.current_readme_row}')
            
            self.current_readme_row += 1
        
        # Add custom metrics from database
        for metric in custom_metrics:
            if custom_start is None:
                custom_start = self.current_readme_row
            
            self._place_metric_row_from_db(ws, metric)
            
            # Track weight cell
            weight = getattr(metric, 'METRIC_WEIGHT', None) or getattr(metric, 'weight', 5)
            if weight:
                self.weight_cells.append(f'M{self.current_readme_row}')
            
            self.current_readme_row += 1
        
        # Merge category cells
        if evergreen_start and custom_start:
            ws.merge_cells(f'B{evergreen_start}:C{custom_start-1}')
            ws[f'B{evergreen_start}'] = "Evergreen Metric"
            ws[f'B{evergreen_start}'].font = FONTS["evergreen_label"]
            ws[f'B{evergreen_start}'].alignment = Alignment(horizontal='center', vertical='center')
        
        if custom_start:
            ws.merge_cells(f'B{custom_start}:C{self.current_readme_row-1}')
            ws[f'B{custom_start}'] = "Customized Metric"
            ws[f'B{custom_start}'].font = FONTS["custom_label"]
            ws[f'B{custom_start}'].alignment = Alignment(horizontal='center', vertical='center')
    
    def _place_metric_row_from_db(self, ws, metric):
        """Place a single metric row from database record."""
        row = self.current_readme_row
        
        # Metric name (D:E merged)
        ws.merge_cells(f'D{row}:E{row}')
        metric_name = getattr(metric, 'METRIC_NAME', 'Unknown Metric')
        ws[f'D{row}'] = metric_name
        ws[f'D{row}'].font = Font(name='Calibri', size=16)
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Definition (F:I merged)
        ws.merge_cells(f'F{row}:I{row}')
        definition = getattr(metric, 'METRIC_DEF', '')
        ws[f'F{row}'] = definition
        ws[f'F{row}'].font = Font(name='Calibri', size=16)
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Notes (J:L merged)
        ws.merge_cells(f'J{row}:L{row}')
        notes = getattr(metric, 'METRIC_NOTES', '')
        ws[f'J{row}'] = notes
        ws[f'J{row}'].font = Font(name='Calibri', size=16)
        ws[f'J{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Weight (M)
        weight = getattr(metric, 'METRIC_WEIGHT', None) or getattr(metric, 'weight', 5)
        try:
            ws[f'M{row}'] = int(weight) if weight else 5
        except (ValueError, TypeError):
            ws[f'M{row}'] = 5  # Default weight
        ws[f'M{row}'].font = Font(name='Calibri', size=16)
        ws[f'M{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row height
        ws.row_dimensions[row].height = 40
        
        # Apply borders
        for col in range(2, 18):  # B to Q
            ws.cell(row=row, column=col).border = BORDERS["thin"]
    
    def _place_metrics_headers(self, ws):
        """Place metric table headers."""
        row = self.current_readme_row
        
        # Headers with merging
        headers = [
            ('B', 'C', 'Category'),
            ('D', 'E', 'Metrics'),
            ('F', 'I', 'Definitions'),
            ('J', 'L', 'Notes'),
            ('M', 'M', 'Weights'),
            ('N', 'O', 'Weight Definition'),
            ('P', 'Q', 'Scoring Definition')
        ]
        
        for start_col, end_col, text in headers:
            if start_col != end_col:
                ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
            ws[f'{start_col}{row}'] = text
            ws[f'{start_col}{row}'].font = FONTS["header_white"]
            ws[f'{start_col}{row}'].fill = COLORS["purple_header"]
            ws[f'{start_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply borders
        for col in range(2, 18):
            ws.cell(row=row, column=col).border = BORDERS["thick"]
        
        ws.row_dimensions[row].height = ROW_HEIGHTS.get("metric_header", 35)
        self.current_readme_row += 1
    
    def _place_stakeholder_section(self, ws):
        """Place stakeholder perspective input area."""
        start_row = self.current_readme_row
        end_row = start_row + 6  # 7 rows for text input
        
        ws.merge_cells(f'B{start_row}:Q{end_row}')
        
        # Pre-fill with stakeholder perspective if provided
        if hasattr(self.state, 'stakeholder_perspective') and self.state.stakeholder_perspective:
            ws[f'B{start_row}'] = self.state.stakeholder_perspective
        
        ws[f'B{start_row}'].font = FONTS.get("normal", Font(name='Calibri', size=14))
        ws[f'B{start_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Apply border
        self._apply_block_border(ws, start_row, end_row, 2, 17)
        
        self.current_readme_row = end_row + 1
    
    def _place_metrics_total(self, ws):
        """Place the metric weightings total sum."""
        row = self.current_readme_row
        
        # Merge cells for total label
        ws.merge_cells(f'J{row}:L{row}')
        ws[f'J{row}'] = "Total:"
        ws[f'J{row}'].font = Font(bold=True, size=16)
        ws[f'J{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Sum formula in M column
        if self.weight_cells:
            ws[f'M{row}'] = f"=SUM({','.join(self.weight_cells)})"
        else:
            ws[f'M{row}'] = 0
        ws[f'M{row}'].font = Font(bold=True, size=16)
        ws[f'M{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply borders
        for col in range(10, 14):  # J to M
            ws.cell(row=row, column=col).border = BORDERS["thick"]
        
        self.current_readme_row += 2
    
    def _place_scoring_definitions(self, ws):
        """Place scoring definitions section."""
        # This can be customized based on your needs
        # For now, just adding a placeholder
        row = self.current_readme_row
        ws[f'B{row}'] = "Scoring Definitions:"
        ws[f'B{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'B{row}:Q{row}')
        
        row += 1
        scoring_text = """
        5 - Excellent: No issues found
        4 - Good: Minor issues that don't affect understanding
        3 - Average: Noticeable issues but still acceptable
        2 - Below Average: Significant issues affecting quality
        1 - Poor: Major issues, needs complete revision
        """
        
        ws[f'B{row}'] = scoring_text
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'B{row}:Q{row+5}')
    
    def _apply_sheet_config(self, ws, config_key: str):
        """Apply sheet configuration like column widths."""
        if config_key in COLUMN_WIDTHS:
            for col, width in COLUMN_WIDTHS[config_key].items():
                ws.column_dimensions[col].width = width
    
    def _apply_block_border(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """Apply border around a block of cells."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    # ============ FORMULA HELPER ============
    def _create_formula_helper(self):
        """Create hidden formula helper sheet."""
        try:
            ws = self.wb.create_sheet("FORMULA_HELPER")
            ws.sheet_state = 'hidden'
            
            # Headers
            headers = ["Evergreen Metric", "Evergreen Weights", "Total Evergreen",
                      "Custom Metric", "Custom Weights", "Total Custom", "Total Combined Weight SUM"]
            
            for i, header in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=header)
                ws.cell(row=1, column=i).font = Font(bold=True)
                ws.cell(row=1, column=i).fill = COLORS["light_blue"]
                ws.cell(row=1, column=i).border = BORDERS["thin"]
            
            # Place weights for selected metrics from database
            evergreen_row = 2
            custom_row = 2
            
            # Get metrics from state
            evergreen_metrics = getattr(self.state, 'evergreen_metrics_db', [])
            custom_metrics = getattr(self.state, 'custom_metrics_db', [])
            
            # Add evergreen metrics
            for metric in evergreen_metrics:
                metric_name = getattr(metric, 'METRIC_NAME', 'Unknown')
                ws[f'A{evergreen_row}'] = metric_name
                
                weight = getattr(metric, 'METRIC_WEIGHT', None) or 5
                try:
                    ws[f'B{evergreen_row}'] = int(weight)
                except (ValueError, TypeError):
                    ws[f'B{evergreen_row}'] = 5
                evergreen_row += 1
            
            # Add custom metrics
            for metric in custom_metrics:
                metric_name = getattr(metric, 'METRIC_NAME', 'Unknown')
                ws[f'D{custom_row}'] = metric_name
                
                weight = getattr(metric, 'METRIC_WEIGHT', None) or 5
                try:
                    ws[f'E{custom_row}'] = int(weight)
                except (ValueError, TypeError):
                    ws[f'E{custom_row}'] = 5
                custom_row += 1
            
            # Sum formulas
            ws['C2'] = f"=SUM(B2:B{max(evergreen_row-1, 2)})"
            ws['F2'] = f"=SUM(E2:E{max(custom_row-1, 2)})"
            ws['G2'] = "=C2+F2"
            
        except Exception as e:
            logger.error(f"Failed to create formula helper sheet: {str(e)}")
            raise
    
    # ============ PART 1, 2, 3 SHEETS ============
    def _create_part1_sheets(self):
        """Create Part 1 evaluation sheets based on number of models."""
        try:
            num_models = int(getattr(self.state, 'num_models', 1))
            
            for i in range(num_models):
                letter = chr(65 + i)  # A, B, C, etc.
                self._create_single_part1_sheet(letter)
                
        except Exception as e:
            logger.error(f"Failed to create Part 1 sheets: {str(e)}")
            raise
    
    def _create_single_part1_sheet(self, model_letter: str):
        """Create a single Part 1 sheet."""
        ws = self.wb.create_sheet(f"PART 1 - MODEL {model_letter}")
        # Simplified implementation - add your full logic here
        ws['A1'] = f"Model {model_letter} Evaluation"
        ws['A1'].font = Font(bold=True, size=16)
    
    def _create_part2_sheet(self):
        """Create Part 2 - Data Analysis sheet."""
        try:
            ws = self.wb.create_sheet("PART 2 - DATA ANALYSIS")
            ws['A1'] = "Data Analysis - To Be Implemented"
            ws['A1'].font = Font(bold=True, size=16)
        except Exception as e:
            logger.error(f"Failed to create Part 2 sheet: {str(e)}")
            raise
    
    def _create_part3_sheet(self):
        """Create Part 3 - Criteria Based Assessment sheet."""
        try:
            ws = self.wb.create_sheet("PART 3 - CRITERIA BASED ASSESS")
            ws['A1'] = "Criteria Based Assessment - To Be Implemented"
            ws['A1'].font = Font(bold=True, size=16)
        except Exception as e:
            logger.error(f"Failed to create Part 3 sheet: {str(e)}")
            raise
    
    def _save_to_bytes(self) -> bytes:
        """Save workbook to bytes."""
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.read()